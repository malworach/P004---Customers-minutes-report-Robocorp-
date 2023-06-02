from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


def Format_Table(xlsx_file):
    # load the workbook
    workbook = load_workbook(xlsx_file)

    # select the worksheet
    worksheet = workbook.active

    # create a list of rows, excluding the header row
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    # sort the rows based on the value in column D
    sorted_rows = sorted(rows, key=lambda row: row[3])

    # overwrite the existing data in the worksheet with the sorted rows
    for row_index, row in enumerate(sorted_rows, start=2):
        for col_index, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index, value=cell_value)

    # set column width
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['C'].width = 40
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['E'].width = 20

    # set first row as headers and format them
    header_style = Font(bold=True, color='FFFFFF')
    fill_style = PatternFill(fill_type='solid', fgColor='0070C0')
    border_style = Border(bottom=Side(border_style='thin', color='000000'))
    for cell in worksheet[1]:
        cell.font = header_style
        cell.fill = fill_style
        cell.border = border_style
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # adjust the height of the first row
    worksheet.row_dimensions[1].height = 20

    # freeze the first row
    worksheet.freeze_panes = 'A2'

    # create the table and apply style
    table = Table(
        displayName='Table1',
        ref=f'A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}')
    style = TableStyleInfo(name='TableStyleMedium2',
                           showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=True,
                           showColumnStripes=False)
    table.tableStyleInfo = style
    table.autoFilter = None  # disable auto-filtering
    worksheet.add_table(table)

    # save the modified workbook
    workbook.save(xlsx_file)