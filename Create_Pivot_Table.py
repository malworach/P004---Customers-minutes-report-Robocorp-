import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side


def Create_Pivot_Table(xlsx_file):
    sheet_name = 'Pivot'

    df = pd.read_excel(xlsx_file)

    # create pivot table
    pivot_table = pd.pivot_table(df,
                                 values='Sum of Session Duration',
                                 index='Process Name',
                                 aggfunc='sum')

    # open existing workbook
    workbook = load_workbook(xlsx_file)

    # add a new worksheet
    worksheet = workbook.create_sheet(sheet_name)

    # write the pivot table to the worksheet
    for r in dataframe_to_rows(pivot_table, index=True, header=True):
        worksheet.append(r)

# Set column widths
    worksheet.column_dimensions['A'].width = 60
    worksheet.column_dimensions['B'].width = 25

    # Set table border style
    table_border = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    # Set header style
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.border = table_border
        #cell.alignment = Alignment(horizontal='center', vertical='center')
        if cell.column_letter == 'B':
            cell.alignment = Alignment(horizontal='right',
                                       vertical='center',
                                       wrap_text=True)
            cell.font = header_font


# Set table style
    for row in worksheet['A2:B{}'.format(pivot_table.shape[0] + 2)]:
        for cell in row:
            cell.border = table_border
            #cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.column_letter == 'A':
                cell.alignment = Alignment(horizontal='left',
                                           vertical='top',
                                           wrap_text=True)
            if cell.row == 2:
                cell.font = header_font

    last_row = worksheet.max_row + 1
    worksheet.cell(row=last_row, column=1).value = 'Grand Total'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=2).value = sum(
        pivot_table['Sum of Session Duration'])
    worksheet.cell(row=last_row, column=2).border = table_border
    worksheet.cell(row=last_row,
                   column=2).alignment = Alignment(horizontal='right',
                                                   vertical='center')
    worksheet.cell(row=last_row, column=2).font = header_font

    workbook.save(xlsx_file)
    workbook.close()